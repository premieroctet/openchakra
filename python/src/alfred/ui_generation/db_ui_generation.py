'''
Created on 31 août 2021

@author: seb
'''
import openpyxl as xl


import argparse
import os
import sys

class FatalException(Exception):
  pass

class DBUiGeneration(object):
    
    COMP='comp'
    
    REQUEST="db.uiconfigurations.update("\
        "{{classname: '{}', type:'{}'}},"\
        "{{$set : {{page: '{}', component: '{}', label: '{}'}}}},"\
        "{{upsert: true}}"\
        ")"
    def __init__(self, xl_path):
      self.wb=xl.load_workbook(xl_path)
      
    def get_column(self, sheet, col_name, mandatory=True):
      first_row=[(c.value or '').lower() for c in next(sheet.iter_rows())]
      try:
        return first_row.index(col_name)
      except Exception as e:
        if (mandatory):
          raise e
        return None
    
    def escape(self, txt):
      return txt.replace("'", "\\'").replace("&", "\\&")
    
    def get_mongo(self, page, component, label, classname, type_):
      return  self.REQUEST.format(*map(self.escape, [classname, type_, page, component, label]))
    
    def extractComponent(self, comp, label):
      first, compName=label.split(' ',1)
      index=int(first[len(self.COMP):])
      if (index-1)>len(comp):
        raise FatalException('Composant de niveau {} sans composant de niveau {}'.format(index, index-1))
      comp=comp[:index-1]+[compName]
      return comp
      
    def export(self):

      for name in self.wb.sheetnames:
        try:
          print('// Handling page {}'.format(name), file=sys.stderr)
          sheet=self.wb[name]
          name=name.replace('Page ', '')
          classnameIdx=self.get_column(sheet, 'classname')
          compTypeIdx=self.get_column(sheet, 'type', False)

          comp=[]
          label=''
          for idx, row in enumerate(list(sheet.iter_rows())[1:]):
            try:
              row=[r.value for r in row]
              col0=row[0]
              if col0:
                if col0.lower().startswith(self.COMP):
                  comp=self.extractComponent(comp, col0)
                  label=''
                elif col0.lower()!='actions':
                  label=col0
                  
              classname=row[classnameIdx]
              compType=row[compTypeIdx] if compTypeIdx!=None else ''
              if classname and classname.lower().replace(' ', '')=='pasducss':
                classname=''
              if compType and not classname:
                raise Exception('Type sans classname')

              if classname:
                if compType:
                  componentType=compType
                else:
                  componentType = classname.split('.')[-1]
                  classname='.'.join(classname.split('.')[:-1])
              if classname:
                if not comp:
                  raise Exception('Classname sans composant')
                req=self.get_mongo(name.capitalize(), ".".join(comp).capitalize(), label, classname, componentType)
                print(req)
                if ('i18N' in str(row[1]).lower()) and not [v for v in row if ('text' in str(v).lower() or 'sample' in str(v).lower())]:
                  raise Exception("Pas de type 'text' pour un label I81N")
            except Exception as ex:
              print('****** Page {}, ligne {}:{}'.format(name, idx+2, ex), file=sys.stderr)
              if isinstance(ex, FatalException):
                print('Exiting', file=sys.stderr)
                sys.exit(0)
        except Exception as ex:
          print('Page {}:{}'.format(name, ex), file=sys.stderr)

      
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Generation script SQL d'import de la configuration UI")
    parser.add_argument('xl_input', default=None, help='fichier Excel de définition des atrtributs')
    args = parser.parse_args()
    
    exp=DBUiGeneration(args.xl_input)
    exp.export()
