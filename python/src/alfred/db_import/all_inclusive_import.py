import sys

from alfred.database.db_access import DBAccess
from alfred.misc.utils import get_items
import openpyxl as xl


class AllInclusiveImport(object):

    def __init__(self, db_name):
        super(object, self).__init__()
        self.db = DBAccess(db_name)
       
    def get_column(self, sheet, col_name, mandatory=True):
      first_row = [(c.value or '').lower() for c in next(sheet.iter_rows())]
      try:
        return first_row.index(col_name)
      except Exception as e:
        if (mandatory):
          raise e
        return None
    
    def get_prestation(self, category_lbl, service_lbl, presta_lbl, filter_lbl, part_access, pro_access):
      part_access = part_access == 'O'
      pro_access = pro_access == 'O'
      cat=self.db.get_item('categories', {'particular_label': category_lbl})
      if cat:
        cat=cat['_id']
      else:
        cat=self.db.insert_document('categories', {'particular_label': category_lbl, 'professional_label': category_lbl})
        
      service=self.db.get_item('services', {'label': service_lbl, 'category': cat})
      if service:
        service = service['_id']
      else:
        service=self.db.insert_document('services', 
            {'label': service_lbl, 'category': cat, 'particular_access': part_access, 'professional_access': pro_access}
        )
      
      if filter_lbl:  
        filter_ = self.db.get_item('filterpresentations', {'label': filter_lbl})
        if filter_:
          filter_ = filter_['_id']
        else:
          filter_=self.db.insert_document('filterpresentations', {'label': filter_lbl})
      else:
        filter_ = None
      
      presta=self.db.get_item('prestations', {'label': presta_lbl, 'service': service, 'filter_presentation': filter_})
      if not presta:
        presta=self.db.insert_document('prestations', 
            {'label': presta_lbl, 'service': service, 'category': cat, 
             'filter_presentation': filter_, 'particular_access': part_access, 'professional_access': pro_access })
      else:
        self.db.update_document('prestations', {'_id': presta['_id'], 'particular_access': part_access, 'professional_access': pro_access})
      return (cat, service,presta)

    def set_billing(self, prestation, billing_lbl):
      if not billing_lbl:
        return
      billings=[self.db.get_item('billings', {'label': lbl})['_id'] for lbl in billing_lbl.split(';')]
      if None in billings:
        raise Exception('Billing inconnu:{}'.format(billing_lbl))
      self.db.update_document('prestations', {'_id': prestation, 'billing': billings})

    def set_equipments(self, service, equipments_lbl):
      if not equipments_lbl:
        return
      equipments=[self.db.get_item('equipment', {'label': lbl})['_id'] for lbl in equipments_lbl.split(';')]
      if None in equipments:
        raise Exception('Equipementinconnu:{}'.format(equipments_lbl))
      self.db.update_document('services', {'_id': service, 'equipments': equipments})

    def set_location(self, service, client, alfred, visio):
      location={'client':client=='O', 'alfred': alfred=='O', 'visio': visio=='O'}
      self.db.update_document('services', {'_id': service, 'location': location})

    def set_pick_tax(self, service, tax):
      self.db.update_document('services', {'_id': service, 'pick_tax': tax=='O'})

    def set_travel_tax(self, service, tax):
      self.db.update_document('services', {'_id': service, 'travel_tax': tax=='O'})

    def set_cesu(self, prestation, cesu_lbl):
      if not cesu_lbl:
        raise Exception('Eligible CESU indéfini')
      self.db.update_document('prestations', {'_id': prestation, 'cesu_eligible': cesu_lbl=='O'})

    def set_job(self, prestation, job_lbl):
      if not job_lbl:
        return
      job=self.db.get_item('jobs', {'label': job_lbl})
      if job:
        job = job['_id']
      else:
        job=self.db.insert_document('jobs', {'label' : job_lbl})
      self.db.update_document('prestations', {'_id': prestation, 'job': job})
      
    def load_defaults(self):
      pass
      self.equipments = self.db.get_items('equipments')

    def clear_db(self):
      return
      for collname in 'availabilities serviceusers services categories prestations bookings reviews chatrooms reviews shops'.split(' '):
        self.db.remove_all_documents(collname, skip_error=True)
      print("Cleared DB")

    def import_file(self, xl_path):
      self.clear_db()
      wb = xl.load_workbook(xl_path)
      sheet=wb['BDD All Inclusive']
      for row in list(sheet.iter_rows())[1:]:
        row=[r.value for r in row]
        print(row)
        if row[0]:
          (_category, service, prestation) = self.get_prestation(row[0], row[1], row[2], None if row[5]=='Aucun' else row[5], *row[13:15])
          self.set_billing(prestation, row[3])
          self.set_cesu(prestation, row[4])
          self.set_equipments(service, row[6])
          self.set_job(prestation, row[10])
          self.set_location(service, *row[7:10])
          self.set_pick_tax(service, row[11])
          self.set_travel_tax(service, row[12])
      
      
if __name__ == '__main__':
    exp = AllInclusiveImport(sys.argv[1])
    exp.import_file(sys.argv[2])
